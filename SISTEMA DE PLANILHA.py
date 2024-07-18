import tkinter
from tkinter import ttk, messagebox, filedialog
import os
import openpyxl
from datetime import datetime
from PIL import Image, ImageTk

#BACK END

def validate_datetime_format(input_string, format_string):
    try:
        datetime.strptime(input_string, format_string)
        return True
    except ValueError:
        return False

def find_last_row(sheet, column):
    max_row = sheet.max_row
    for row in range(max_row, 1, -1):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value is not None and not sheet.row_dimensions[row].hidden:
            return row + 1
    return 1

def on_combobox_change(event, combobox, values_list):
    typed_text = combobox.get().strip()
    
    if typed_text == "":
        combobox['values'] = values_list
    else:
        filtered_values = [value for value in values_list if value.lower().startswith(typed_text.lower())]
        combobox['values'] = filtered_values

    # Manter o texto digitado pelo usuário visível no combobox
    combobox.icursor(tkinter.END)  # Colocar o cursor no final do texto

    # Verificar se há apenas uma opção na lista filtrada e definir no combobox
    if len(filtered_values) == 1:
        combobox.set(filtered_values[0])
        combobox.icursor(tkinter.END)  # Colocar o cursor no final do texto após definir

    # Selecionar todo o texto para mostrar a seleção
    combobox.select_range(0, tkinter.END)

def choose_excel_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    if filepath:
        excel_file_entry.delete(0, tkinter.END)
        excel_file_entry.insert(0, filepath)

def validate_time_format(input_string):
    try:
        datetime.strptime(input_string, "%H:%M")
        return True
    except ValueError:
        return False

def load_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return [cell.value for cell in sheet['A'] if cell.value]

def validate_and_enter_data():
    accepted = accept_var.get()
    filepath = excel_file_entry.get()
    
    if accepted == "Accepted":
        if not (cliente_combobox.get() and coordenadas_entry.get() and cidade_entry.get() and estado_combobox.get() and
                data_consulta_entry.get() and data_disponibilidade_entry.get() and data_aprovacao_entry.get() and
                motivo_nao_atendimento_combobox.get() and operador_plantao_combobox.get() and supervisor_combobox.get()):
            messagebox.showerror("Erro", "Por favor, preencha todos os campos obrigatórios.")
            return
        
        cliente = cliente_combobox.get()
        coordenadas = coordenadas_entry.get()
        cidade = cidade_entry.get()
        estado = estado_combobox.get()
        atendido = atendido_var.get()
        data_consulta = data_consulta_entry.get()
        data_disponibilidade = data_disponibilidade_entry.get()
        previsao_deslocamento = previsao_deslocamento_entry.get()
        data_aprovacao = data_aprovacao_entry.get()
        motivo_nao_atendimento = motivo_nao_atendimento_combobox.get()
        operador_plantao = operador_plantao_combobox.get()
        supervisor = supervisor_combobox.get()
        
        if data_consulta and not validate_datetime_format(data_consulta, "%d/%m/%Y %H:%M"):
            messagebox.showerror("Erro", "Formato inválido para Data/Hora da Consulta. Utilize dia/mês/ano hora:minutos.")
            return
        
        if data_disponibilidade and not validate_datetime_format(data_disponibilidade, "%d/%m/%Y %H:%M"):
            messagebox.showerror("Erro", "Formato inválido para Data/Hora Disponibilidade Apoio. Utilize dia/mês/ano hora:minutos.")
            return
        
        if data_aprovacao and not validate_datetime_format(data_aprovacao, "%d/%m/%Y %H:%M"):
            messagebox.showerror("Erro", "Formato inválido para Data/Hora Aprovação/QTA. Utilize dia/mês/ano hora:minutos")
            return
        
        if not filepath:
            messagebox.showerror("Erro", "Por favor, selecione um arquivo Excel para salvar os dados.")
            return
        
        if previsao_deslocamento and not validate_time_format(previsao_deslocamento):
            messagebox.showerror("Erro", "Formato inválido para Previsão de Deslocamento. Utilize hora:minutos.")
            return
        elif previsao_deslocamento:
            previsao_deslocamento = datetime.strptime(previsao_deslocamento, "%H:%M").strftime("%H:%M")
        
        if data_consulta:
            data_consulta = datetime.strptime(data_consulta, "%d/%m/%Y %H:%M").strftime("%d/%m/%Y %H:%M")
        if data_disponibilidade:
            data_disponibilidade = datetime.strptime(data_disponibilidade, "%d/%m/%Y %H:%M").strftime("%d/%m/%Y %H:%M")
        if data_aprovacao:
            data_aprovacao = datetime.strptime(data_aprovacao, "%d/%m/%Y %H:%M").strftime("%d/%m/%Y %H:%M")
        
        print("Cliente:", cliente)
        print("Coordenadas:", coordenadas)
        print("Cidade:", cidade)
        print("Estado:", estado)
        print("Atendido:", atendido)
        print("Data/Hora da Consulta:", data_consulta)
        print("Data/Hora Disponibilidade Apoio:", data_disponibilidade)
        print("Previsão de Deslocamento:", previsao_deslocamento)
        print("Data/Hora Aprovação/QTA:", data_aprovacao)
        print("Motivo do Não Atendimento:", motivo_nao_atendimento)
        print("Operador de Plantão:", operador_plantao)
        print("Supervisor:", supervisor)
        print("Planilha de Destino:", filepath)
        print("------------------------------------------")
        
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["Cliente", "Coordenadas", "Cidade", "Estado", "Atendido?",
                       "Data/Hora da Consulta", "Data/Hora disponibilidade apoio",
                       "Previsão de Deslocamento", "Data/Hora Aprovação/QTA",
                       "Motivo do Não Atendimento", "Operador de Plantão", "Supervisor"]
            sheet.append(heading)
            workbook.save(filepath)
            
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        last_row = find_last_row(sheet, column=1)
        
        sheet.append([cliente, coordenadas, cidade, estado, atendido,
                      data_consulta, data_disponibilidade, previsao_deslocamento,
                      data_aprovacao, motivo_nao_atendimento, operador_plantao, supervisor])
        
        workbook.save(filepath)
        
        clear_fields()
        accept_var.set("Not Accepted")
        
    else:
        messagebox.showwarning(title="Erro", message="Por favor, aceite os termos para enviar os dados.")

def clear_fields():
    cliente_combobox.set("")
    coordenadas_entry.delete(0, tkinter.END)
    cidade_entry.delete(0, tkinter.END)
    estado_combobox.set("")
    atendido_var.set("")
    data_consulta_entry.delete(0, tkinter.END)
    data_disponibilidade_entry.delete(0, tkinter.END)
    previsao_deslocamento_entry.delete(0, tkinter.END)
    data_aprovacao_entry.delete(0, tkinter.END)
    motivo_nao_atendimento_combobox.set("")
    operador_plantao_combobox.set("")
    supervisor_combobox.set("")




# INICIO DA INTERFACE (FRONT END)

window = tkinter.Tk()
window.title("Formulário de Consultas RNS")
window.configure(bg="#4169E1")


logo_path = "rnslogo.png"  # Substitua pelo caminho do seu arquivo de logo
logo_image = tkinter.PhotoImage(file=logo_path)
logo_label = tkinter.Label(window, image=logo_image, bg='#4169E1')
logo_label.place(x=0, y=0)  # Posicionando no canto superior esquerdo


frame = tkinter.Frame(window, bg='#4169E1')
frame.pack(pady=20)


frame = tkinter.Frame(window, bg='#4169E1')
frame.pack()

# Dados DA CONSULTA
consulta_frame = tkinter.LabelFrame(frame, text="DADOS DA CONSULTA", bg="#4169E1", font=("Calibri", 14, "bold"), fg="white")
consulta_frame.grid(row=0, column=0, padx=20, pady=10)


# Carregar dados de arquivos Excel
clientes = load_data('clientes.xlsx')
operadores = load_data('operadores.xlsx')
supervisores = load_data('supervisores.xlsx')


cliente_combobox = ttk.Combobox(consulta_frame, values=clientes, style="Custom.TCombobox", font=("Calibri", 12, "bold"))
cliente_label = tkinter.Label(consulta_frame, text="Cliente", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
cliente_label.grid(row=0, column=0, padx=10, pady=5)
cliente_combobox.grid(row=0, column=1, padx=10, pady=5)
cliente_combobox.bind("<KeyRelease>", lambda event: on_combobox_change(event, cliente_combobox, clientes))


coordenadas_label = tkinter.Label(consulta_frame, text="Coordenadas", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
coordenadas_label.grid(row=1, column=0, padx=10, pady=5)
coordenadas_entry = tkinter.Entry(consulta_frame, font=("Calibri", 12))
coordenadas_entry.grid(row=1, column=1, padx=10, pady=5)


cidade_label = tkinter.Label(consulta_frame, text="Cidade", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
cidade_label.grid(row=2, column=0, padx=10, pady=5)
cidade_entry = tkinter.Entry(consulta_frame, font=("Calibri", 12))
cidade_entry.grid(row=2, column=1, padx=10, pady=5)


estados_brasil = [
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA",
    "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"
]
estado_label = tkinter.Label(consulta_frame, text="Estado", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
estado_label.grid(row=3, column=0, padx=10, pady=5)
estado_combobox = ttk.Combobox(consulta_frame, values=estados_brasil, font=("Calibri", 12))
estado_combobox.grid(row=3, column=1, padx=10, pady=5)
estado_combobox.bind("<KeyRelease>", lambda event: on_combobox_change(event, estado_combobox, estados_brasil))


# Rótulo e botões de opção para "Atendido?"
atendido_label = tkinter.Label(consulta_frame, text="Atendido?", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
atendido_label.grid(row=4, column=0, padx=10, pady=5)

atendido_var = tkinter.StringVar(value="Não")
atendido_sim = tkinter.Radiobutton(consulta_frame, text="Sim", variable=atendido_var, value="Sim", font=("Calibri", 12))
atendido_sim.grid(row=4, column=1, padx=10, pady=5)

atendido_nao = tkinter.Radiobutton(consulta_frame, text="Não", variable=atendido_var, value="Não", font=("Calibri", 12))
atendido_nao.grid(row=4, column=2, padx=10, pady=5)


data_consulta_label = tkinter.Label(consulta_frame, text="Data/Hora da Consulta (dia/mês/ano hora:minutos)", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
data_consulta_label.grid(row=5, column=0, padx=10, pady=5)
data_consulta_entry = tkinter.Entry(consulta_frame, font=("Calibri", 12))
data_consulta_entry.grid(row=5, column=1, padx=10, pady=5)


data_disponibilidade_label = tkinter.Label(consulta_frame, text="Data/Hora Disponibilidade Apoio (dia/mês/ano hora:minutos)", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
data_disponibilidade_label.grid(row=6, column=0, padx=10, pady=5)
data_disponibilidade_entry = tkinter.Entry(consulta_frame, font=("Calibri", 12))
data_disponibilidade_entry.grid(row=6, column=1, padx=10, pady=5)


previsao_deslocamento_label = tkinter.Label(consulta_frame, text="Previsão de Deslocamento (hora:minutos)", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
previsao_deslocamento_label.grid(row=7, column=0, padx=10, pady=5)
previsao_deslocamento_entry = tkinter.Entry(consulta_frame, font=("Calibri", 12))
previsao_deslocamento_entry.grid(row=7, column=1, padx=10, pady=5)


data_aprovacao_label = tkinter.Label(consulta_frame, text="Data/Hora Aprovação/QTA (dia/mês/ano hora:minutos)", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
data_aprovacao_label.grid(row=8, column=0, padx=10, pady=5)
data_aprovacao_entry = tkinter.Entry(consulta_frame, font=("Calibri", 12))
data_aprovacao_entry.grid(row=8, column=1, padx=10, pady=5)


motivo_nao_atendimento_label = tkinter.Label(consulta_frame, text="Motivo do Não Atendimento", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
motivo_nao_atendimento_label.grid(row=9, column=0, padx=10, pady=5)
motivo_nao_atendimento_combobox = ttk.Combobox(consulta_frame, values=[
    "Cliente solicitou cancelamento sem justificativa",
    "Atendido",
    "Sem apoio na região",
    "Cliente cancelou antes de 10 minutos",
    "Sem equipe disponível",
    "Conseguiu contato com o condutor",
    "Apoio muito distante",
    "Veiculo voltou a posicionar",
    "Central não visualizou a tempo",
    "Região de risco"
], state="readonly", font=("Calibri", 12))
motivo_nao_atendimento_combobox.grid(row=9, column=1, padx=10, pady=5)


# Dados DO ATENDIMENTO
atendimento_frame = tkinter.LabelFrame(frame, text="SELECIONAR A PLANILHA", bg="#4169E1", font=("Calibri", 14, "bold"), fg="white")
atendimento_frame.grid(row=1, column=0, padx=20, pady=10)


operador_plantao_label = tkinter.Label(consulta_frame, text="Operador de Plantão", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
operador_plantao_label.grid(row=10, column=0, padx=10, pady=5)
operador_plantao_combobox = ttk.Combobox(consulta_frame, values=operadores, font=("Calibri", 12))
operador_plantao_combobox.grid(row=10, column=1, padx=10, pady=5)
operador_plantao_combobox.bind("<KeyRelease>", lambda event: on_combobox_change(event, operador_plantao_combobox, operadores))


supervisor_label = tkinter.Label(consulta_frame, text="Supervisor", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
supervisor_label.grid(row=11, column=0, padx=10, pady=5)
supervisor_combobox = ttk.Combobox(consulta_frame, values=supervisores, font=("Calibri", 12))
supervisor_combobox.grid(row=11, column=1, padx=10, pady=5)
supervisor_combobox.bind("<KeyRelease>", lambda event: on_combobox_change(event, supervisor_combobox, supervisores))


aceitar_termos_label = tkinter.Label(atendimento_frame, bg="#4169E1", fg="white", text="Preenchido corretamente?", font=("Calibri", 12, "bold"))
accept_var = tkinter.StringVar(value="Not Accepted")
aceitar_termos_checkbutton = tkinter.Checkbutton(atendimento_frame, text="Sim!", variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
aceitar_termos_label.grid(row=2, column=0)
aceitar_termos_checkbutton.grid(row=2, column=1)


excel_file_label = tkinter.Label(atendimento_frame, text="Planilha de Consulta", bg="#4169E1", fg="white", font=("Calibri", 12, "bold"))
excel_file_entry = tkinter.Entry(atendimento_frame, width=50)
excel_file_button = tkinter.Button(atendimento_frame, text="Escolher arquivo", command=choose_excel_file, bg="white", fg="#4169E1", font=("Calibri", 12, "bold"))
excel_file_label.grid(row=3, column=0)
excel_file_entry.grid(row=3, column=1)
excel_file_button.grid(row=3, column=2)


enviar_button = tkinter.Button(frame, text="Enviar Consulta", bg="white", fg="#4169E1", font=("Calibri", 12, "bold"), command=validate_and_enter_data)
enviar_button.grid(row=2, column=0, pady=10)


window.mainloop()